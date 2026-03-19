"""
XLSX Connector
--------------
Reads all sheets in an Excel workbook using openpyxl.
Each sheet is converted to a markdown-style text table and returned
as a separate LangChain Document.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import List

import openpyxl
from langchain_core.documents import Document

from logger.custom_logger import CustomLogger
from exception.custom_exception import DocumentPortalException

log = CustomLogger().get_logger(__name__)


def _sheet_to_markdown(sheet) -> str:
    """
    Convert an openpyxl worksheet into a pipe-delimited markdown-style table string.
    Skips fully empty rows.
    """
    rows: List[str] = []
    header_written = False

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [str(cell) if cell is not None else "" for cell in row]

        # Skip rows that are entirely empty
        if all(c == "" for c in cells):
            log.debug(
                "XLSX connector: skipping empty row",
                sheet=sheet.title,
                row=row_idx,
            )
            continue

        row_str = "| " + " | ".join(cells) + " |"
        rows.append(row_str)

        # Insert a markdown header separator after the first non-empty row
        if not header_written:
            separator = "| " + " | ".join(["---"] * len(cells)) + " |"
            rows.append(separator)
            header_written = True

    return "\n".join(rows)


def load_xlsx(path: Path) -> List[Document]:
    """
    Load an .xlsx file and return one Document per sheet.

    Args:
        path: Absolute path to the .xlsx file.

    Returns:
        List of Document objects with sheet data formatted as a markdown table.
    """
    log.info("XLSX connector: starting load", file=str(path))
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        docs: List[Document] = []

        log.info(
            "XLSX connector: workbook opened",
            file=str(path),
            sheet_count=len(wb.sheetnames),
            sheets=wb.sheetnames,
        )

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            markdown_table = _sheet_to_markdown(sheet)

            if not markdown_table.strip():
                log.warning(
                    "XLSX connector: sheet is empty — skipping",
                    file=str(path),
                    sheet=sheet_name,
                )
                continue

            content = f"## Sheet: {sheet_name}\n\n{markdown_table}"
            docs.append(
                Document(
                    page_content=content,
                    metadata={
                        "source": str(path),
                        "file_type": ".xlsx",
                        "sheet_name": sheet_name,
                    },
                )
            )
            log.info(
                "XLSX connector: sheet loaded",
                file=str(path),
                sheet=sheet_name,
                rows=sheet.max_row,
                cols=sheet.max_column,
            )

        wb.close()
        log.info(
            "XLSX connector: load complete",
            file=str(path),
            documents_created=len(docs),
        )
        return docs

    except Exception as e:
        log.error("XLSX connector: failed to load file", file=str(path), error=str(e))
        raise DocumentPortalException(f"Failed to load XLSX file: {path.name}", e) from e
